import streamlit as st
import pdfplumber
import re
import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="Extractor de Piezas GNP", page_icon="🔧", layout="centered")
st.title("🔧 Extractor de Piezas Sustituidas")
st.markdown("**Valuaciones GNP / Audatex** — Sube tu PDF y descarga el Excel al instante.")
st.divider()

def extraer_numero_orden(nombre_archivo):
    """Extrae el número de orden del nombre del archivo. Ej: '3278-...' -> '3278'"""
    m = re.match(r'^(\d+)', Path(nombre_archivo).stem)
    return m.group(1) if m else Path(nombre_archivo).stem

def extraer_texto(contenido_bytes):
    texto_completo = []
    with pdfplumber.open(io.BytesIO(contenido_bytes)) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if texto:
                texto_completo.append(texto)
    return "\n".join(texto_completo)

def parsear_piezas(texto):
    lineas = texto.splitlines()

    inicio = None
    for i, linea in enumerate(lineas):
        if "PIEZAS SUSTITUIDAS" in linea.upper():
            inicio = i + 1
            break
    if inicio is None:
        return []

    fin = len(lineas)
    for i in range(inicio, len(lineas)):
        if re.search(r"ahorro|sub\s*total|total\s*piezas", lineas[i], re.IGNORECASE):
            fin = i
            break

    piezas = []
    for linea in lineas[inicio:fin]:
        linea = linea.strip()
        if not linea:
            continue

        m_precio = re.match(r'^([\$][\d,]+\.\d{2})[\*A-Za-z]?\s+', linea)
        if not m_precio:
            continue

        try:
            precio_num = float(m_precio.group(1).replace('$', '').replace(',', ''))
        except:
            precio_num = 0.0

        tokens = linea[m_precio.end():].split()
        if len(tokens) < 4:
            continue

        descripcion = ' '.join(tokens[1:-2]).strip()
        if not descripcion:
            continue
        if re.match(r'^(precio|referencia|descripci)', descripcion, re.IGNORECASE):
            continue

        piezas.append({"precio": precio_num, "descripcion": descripcion})

    return piezas

def generar_excel(piezas, nombre_archivo):
    numero_orden = extraer_numero_orden(nombre_archivo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Piezas Sustituidas"

    azul  = "1F4E79"
    claro = "D9E1F2"
    fill_hdr  = PatternFill("solid", start_color=azul,     end_color=azul)
    fill_alt  = PatternFill("solid", start_color=claro,    end_color=claro)
    fill_blco = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    fill_tot  = PatternFill("solid", start_color=azul,     end_color=azul)
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # Título
    ws.merge_cells("A1:C1")
    ws["A1"] = f"PIEZAS SUSTITUIDAS  |  Orden {numero_orden}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color=azul)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Encabezados — 3 columnas: No. Orden | Precio | Descripción
    headers = [("A", "No. Orden"), ("B", "Precio"), ("C", "Descripción")]
    for col, txt in headers:
        c = ws[f"{col}2"]
        c.value     = txt
        c.fill      = fill_hdr
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borde
    ws.row_dimensions[2].height = 20

    # Datos
    for i, p in enumerate(piezas, start=3):
        fill = fill_alt if i % 2 == 0 else fill_blco

        # Columna A: No. Orden
        co = ws.cell(row=i, column=1, value=numero_orden)
        co.fill      = fill
        co.font      = Font(name="Arial", size=10)
        co.border    = borde
        co.alignment = Alignment(horizontal="center", vertical="center")

        # Columna B: Precio (número)
        cp = ws.cell(row=i, column=2, value=p["precio"])
        cp.number_format = '"$"#,##0.00'
        cp.fill      = fill
        cp.font      = Font(name="Arial", size=10)
        cp.border    = borde
        cp.alignment = Alignment(horizontal="right", vertical="center")

        # Columna C: Descripción
        cd = ws.cell(row=i, column=3, value=p["descripcion"])
        cd.fill      = fill
        cd.font      = Font(name="Arial", size=10)
        cd.border    = borde
        cd.alignment = Alignment(horizontal="left", vertical="center")

    # Fila de TOTAL
    ft = len(piezas) + 3
    fila_ini = 3
    fila_fin = len(piezas) + 2


    ct_label = ws.cell(row=ft, column=1, value="TOTAL")
    ct_label.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ct_label.fill      = fill_tot
    ct_label.alignment = Alignment(horizontal="center", vertical="center")
    ct_label.border    = borde

    ct_sum = ws.cell(row=ft, column=2, value=f"=SUM(B{fila_ini}:B{fila_fin})")
    ct_sum.number_format = '"$"#,##0.00'
    ct_sum.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ct_sum.fill      = fill_tot
    ct_sum.alignment = Alignment(horizontal="right", vertical="center")
    ct_sum.border    = borde

    ct_desc = ws.cell(row=ft, column=3, value="")
    ct_desc.fill   = fill_tot
    ct_desc.border = borde

    ws.row_dimensions[ft].height = 20

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Interfaz ─────────────────────────────────────────────────
pdf_file = st.file_uploader(
    "📄 Sube tu valuación en PDF",
    type=["pdf"],
    help="Archivos PDF de valuaciones GNP / Audatex"
)

if pdf_file is not None:
    numero_orden = extraer_numero_orden(pdf_file.name)

    with st.spinner("Procesando PDF..."):
        contenido = pdf_file.read()
        texto     = extraer_texto(contenido)
        piezas    = parsear_piezas(texto)

    if not piezas:
        st.error("❌ No se encontró la sección 'PIEZAS SUSTITUIDAS'. Verifica que sea una valuación GNP/Audatex.")
    else:
        total = sum(p["precio"] for p in piezas)
        st.success(f"✅ Orden **{numero_orden}** — {len(piezas)} piezas  |  Total: ${total:,.2f}")
        st.subheader("Piezas sustituidas")
        st.dataframe(
            {
                "No. Orden":   [numero_orden] * len(piezas),
                "Precio":      [f"${p['precio']:,.2f}" for p in piezas],
                "Descripción": [p["descripcion"] for p in piezas],
            },
            use_container_width=True,
            hide_index=True
        )
        excel_buf    = generar_excel(piezas, pdf_file.name)
        nombre_excel = f"{numero_orden}_piezas.xlsx"
        st.download_button(
            label="📥 Descargar Excel",
            data=excel_buf,
            file_name=nombre_excel,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

st.divider()
st.caption("Vanguardia Body & Paint — Extractor de valuaciones GNP")

